import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
# print type(wb)


# wb.get_sheet_names()
#
# sheet = wb.get_sheet_by_name('Sheet3')
# print sheet
#
# print type(sheet)
# print sheet.title


sheet = wb.get_sheet_by_name('Sheet1')
# print sheet['A1']
#
# print sheet['A1'].value
#
# c = sheet['B1']
# print c.value
#
# print 'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
#
# print 'Cell ' + c.coordinate + ' is ' + c.value
#
# print sheet['C1'].value


# print sheet.cell(row=1, column=2)
#
# print sheet.cell(row=1, column=2).value
#
# for i in range(1, 8, 2):
#     print i, sheet.cell(row=i, column=2).value


# print sheet.max_row
# print sheet.max_column


# print sheet['A1':'C3']
#
# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print cellObj.coordinate, cellObj.value
#     print('--- END OF ROW ---')


# print sheet.columns[1]
# for cellObj in sheet.columns[1]:
#     print cellObj.value


# print sheet.title
#
# sheet.title = 'fastcampus'
# wb.save('example_copy.xlsx')

